import re
from sys import exit, argv
import getopt
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Border, Side, Alignment, Protection, PatternFill, Color, Font, colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from lxml import etree
import datetime
from os import path, mkdir
import shutil
import pandas as pd

# Global constants for the date and spreadsheet formatting options
DATE = datetime.datetime.today()
border = Border(left=Side(border_style='thin'), # basic black cell border
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
gray_border = Border(left=Side(border_style='double'), # mirrors excel's "Check Cell" built-in style border
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
my_bad = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # cell style config that mirrors excel "Bad" built-in
bad_font = Font(color='9C0006')
my_neutral = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # cell style config that mirrors excel "Neutral" built-in
neutral_font = Font(color='9C5700')
my_good = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # cell style config that mirrors excel "Good" built-in
good_font = Font(color='006100')
my_check = PatternFill(start_color='A5A5A5', end_color='A5A5A5', fill_type='solid') # cell style config that mirrors excel "Check Cell" built-in
check_font = Font(color='FFFFFF', bold=True)
data_val = DataValidation(type="list", formula1='=statuses!$A$2:$A$22', allow_blank=True) # data validation object for limiitng status column values
vuln_name_style = NamedStyle(name="vuln_name_style") # global text alignment and wrapping style specs
vuln_name_style.alignment = Alignment(horizontal='left',
                                      vertical='center',
                                      wrapText=True)
the_rest_style = NamedStyle(name="the_rest_style")
the_rest_style.alignment = Alignment(horizontal='center',
                                     vertical='center',
                                     wrapText=True)
vuln_name_style.border = border # applies the thin black border to all custom cell styles
the_rest_style.border = border

# Takes an error description and exits the program - used during input validation
def _Err_Exit (error_text):
    print(error_text)
    exit()

# Display commandline help text
def _Opt_Help ():
    help = r"""For CLI usage, provide an option of 1 through 5; otherwise, you'll be prompted to use the interactive prompt.
             Example usage:
                Import .nessus file into spreadsheet:
                    nessus-vuln-analysis.py -2 -n \"C:\Users\Me\Report.nessus\" -s \"C:\Users\Me\Analysis_Spreadsheet.xlsx\"

                -h : print this help text
                -i : follow the script's interactive prompt
                -1 : generate a fresh analysis spreadsheet and exit; optional arguments for new filepath and sheet names
                -2 : import a .nessus file into an existing analysis spreadsheet and exit; optional arguments for paths to nessus file and spreadsheet
                -3 : add a new sheet(s) to an existing analysis spreadsheet and exit; optional argument for sheet name(s)
                -4 : generate a remediation report and exit; optional arguments for sheet name and month number
                -5 : transition to a new spreadsheet, saving it in the same directory as the old one, and exit; optional arguments for old spreadsheet path
                -n : the path to a .nessus report file; include the extension!
                -s : the path to an analysis spreadsheet compatible with this script; include the extension!
                -t : provide a sheet name or list of sheet names (comma-separated, no spaces!) to pass into functions that require them
                -m : provide a number that represents a month; the month number associations are as follows:
                         Jan : 1
                         Feb : 2
                         Mar : 3
                         Apr : 4
                         May : 5
                         Jun : 6
                         Jul : 7
                         Aug : 8
                         Sep : 9
                         Oct : 10
                         Nov : 11
                         Dec : 12"""
    print(help)

# Takes string instructions and a flag for whether or not an extension is expected and exits if too many failed attempts occur or some other issue occurs
def _Check_Path (p, opt):
    count = 0 # counts errors
    while True:
    #     try:
    #         p = str(input(instruct+': '))
    #     except:
    #         "\nUnexpected error. Try again...\n"
        if count == 3:
            _Err_Exit('\nYou seem to be having trouble. Confirm your desired path and come back later.\nExiting...')
        elif opt == 'v': # v option checks if given path / file combo is valid as a new spreadsheet
            f = p.split(".")
            dr = p.split("\\")
            dr2 = "\\".join(dr[:-2])
            if path.isdir(dr2):
                if f[-1] == 'xlsx':
                    break
                else:
                    p = p+'.xlsx'
                    break
            else:
                count+=1
                print('\nInvalid directory selected. Try again...\n')
                continue
        elif opt == 'n': # n option indicates path to .nessus file, checks for valid file based on extension
            if path.isfile(p) and p.split(".")[-1] == "nessus":
                break
            else:
                count+=1
                print('\nThat\'s not a valid nessus file. Try again...\n')
                continue
        elif opt == 'f': # f option indicates path directly to file, checks simply for valid path
            if path.isfile(p):
                break
            else:
                count+=1
                print('\nInvalid path to file. Try again...\n')
                continue
        elif opt == 'd': # d option indicates path to a directory, checks simply for valid path
            if path.isdir(p):
                break
            else:
                count+=1
                print('\nInvalid path to directory. Try again...\n')
                continue
        elif opt == 'x': # x option indicates an excel file that has been generated by this Python program
            print(p.split('.')[-1])
            if p.split('.')[-1] == "xlsx" and path.isdir('\\'.join(p.split('\\')[:-2])): # isolate/check the extension and check validity of target dir
                wb = load_workbook(p)
                sheetlist = []
                for sheet in wb.sheetnames:
                    sheetlist.append(sheet)
                if "statuses" in sheetlist and "columns" in sheetlist: # determines if the excel file has been generated by this program by checking for the existance of specific sheetnames
                    wb.close()
                    break
                else:
                    wb.close()
                    count+=1
                    print("\nThis does not appear to be a compatible workbook. Please provide a workbook that has been generated by this program.\n")
                    continue
            else:
                count+=1
                print('\nInvalid path for writing. Make sure the full path is correct, that the extension is .xlsx, and that the target is a workbook generated originally by this Python tool...\n')
                continue
    return p

# Takes a string input sheet name and checks to see if it's valid within the context of a valid vuln mgmt workbook
def _Check_Sheet (sheet, wb):
    count = 0
    while True:
        if count == 3:
            _Err_Exit('\nYou seem to be having trouble. Confirm your desired sheet\'s name and come back later.\nExiting...')
        elif sheet != 'statuses' and sheet != 'columns':
            try:
                ws = wb[sheet]
                break
            except:
                count+=1
                print("Error in initializing worksheet object. Try again...")
                continue
        else:
            count+=1
            print('\nSheetname cannot be modified. Try again...\n')
            continue
    return sheet

# Takes a path to a new file and makes sure it's a valid directory, exiting if it isn't
def _Check_Opt_Path (opt_path):
    dr = opt_path.split("\\")
    dr2 = "\\".join(dr[:-2])
    if path.isdir(dr2):
        what = 1
    else:
        _Err_Exit("The path you provided is bad.\n")

# Backs up the analysis spreadsheet to a local directory
def _Backup (existing_spreadsheet):
    if path.isdir('\\'.join(existing_spreadsheet.split('\\')[:-1])+'\\Vulnerability Analysis Backups'):
        print("Now saving to backup to Vulnerability Analysis Backups directory...")
        shutil.copyfile(existing_spreadsheet, '\\'.join(existing_spreadsheet.split('\\')[:-1])+'\\Vulnerability Analysis Backups\\'+existing_spreadsheet.split('\\')[-1]+DATE.strftime(" %Y_%m_%d %H%M%S")+'.bak')
    else:
        bak = input("No dedicated backup directory specified in this running location. Would you like to create one? y|n: ")
        if bak == 'y':
            print("Creating backup directory and saving the backup workbook to it...")
            mkdir('\\'.join(existing_spreadsheet.split('\\')[:-1])+'\\\\Vulnerability Analysis Backups')
            shutil.copyfile(existing_spreadsheet, '\\'.join(existing_spreadsheet.split('\\')[:-1])+'\\Vulnerability Analysis Backups\\'+existing_spreadsheet.split('\\')[-1]+DATE.strftime(" %Y_%m_%d %H%M%S")+'.bak')
        else:
            print("Backup file is being saved to the current working directory...")
            shutil.copyfile(existing_spreadsheet, ''.join(existing_spreadsheet.split('\\')[:-2])+existing_spreadsheet.split('\\')[:-1]+DATE.strftime(" %Y_%m_%d %H%M%S")+'.bak')

# Takes the Nessus XML report and generates a dictionary
def _Parse_Nessus(report_path):
    client = ""
    report_dict = dict()
    host_params = ["HOST_START",
                   "mac-address",
                   "netbios-name",
                   "host-rdns",
                   "operating-system",
                   "host-ip"]
    vuln_params = ["pluginName",
                   "pluginID",
                   "port",
                   "svc_name",
                   "severity",
                   "synopsis",
                   "solution",
                   "plugin_output"]
    cred_fail_plugins = ["117886",
                         "21745",
                         "110385"]

    # Open and read the XML report data into the program
    f = open(report_path, 'r')
    xml_content = f.read()
    f.close()

    parz = etree.XMLParser(huge_tree=True) # initialize the parser object
    root = etree.fromstring(text=xml_content, parser=parz) # parse the XML content

    # Iterate over the parsed XML object and generate a dictionary that contains useful values.
    for block in root:
      if block.tag == "Report":
          client = block.attrib['name'].split(" ", 1)[0] # grabs the client acronym from the scan name
          for ReportHost in block:
              props_dict = dict() # dict for holding host properties
              vulns_dict = dict() # dict for holding individual vulnerability dicts
              for ReportItem in ReportHost:
                  if ReportItem.tag == "HostProperties": # assemble host properties
                      for prop in ReportItem:
                          if prop.attrib['name'] in host_params:
                              if prop.attrib['name'] == "mac-address" and len(prop.text) > 17: # if property is mac, sorts them so that they are the same order every run
                                  macs = prop.text.split("\n", 100)
                                  macs.sort()
                                  final_macs = '\n'.join(macs)
                                  props_dict[prop.attrib['name']] = final_macs
                              else:
                                  props_dict[prop.attrib['name']] = prop.text
                  else: # assemble vuln details
                      vuln_dict = dict()
                      for attr in ReportItem.attrib:
                          if attr in vuln_params:
                              vuln_dict[attr] = ReportItem.attrib[attr]
                      for param in ReportItem:
                          if param.tag in vuln_params:
                              vuln_dict[param.tag] = param.text
                      vulns_dict[ReportItem.attrib['pluginID']] = vuln_dict
                  props_dict['vulns'] = vulns_dict
              report_dict[ReportHost.attrib['name']] = props_dict

      # Determine credentialed scan status of each host and create a dictionary key for each
      for host in report_dict:
          fail_count = 0
          for prop in report_dict[host]:
              if prop == "vulns":
                  for plugin in report_dict[host][prop]:
                      if plugin in cred_fail_plugins:
                          fail_count+=1
          if fail_count == 0:
              report_dict[host]["auth"] = 's'
          else:
              report_dict[host]["auth"] = 'f'

    return report_dict, client

# Setting column styles is needed several times throughout the program's functions
def _Set_Col_Styles (ws):
    for cell in ws['A']:
        cell.style = 'vuln_name_style'
    for cell in ws['J']:
        cell.style = 'vuln_name_style'
    for column in ws['B:I']:
        for cell in column:
            cell.style = 'the_rest_style'
    for column in ws['K:V']:
        for cell in column:
            cell.style = 'the_rest_style'
    return ws

# Setting column dimensions is needed several times throughout the program's functions
def _Set_Col_Widths (ws):
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 8
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 25
    ws.column_dimensions['Q'].width = 25
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['T'].width = 18
    ws.column_dimensions['U'].width = 12
    ws.column_dimensions['V'].width = 18
    return ws

# Setting cell format based on the row's 'Status' field is needed several times throughout the program's functions
def _Set_Row_Format (ws):
    for row in ws.iter_rows():
        if row[18].value == '':
            row[18].value = 'Pending Analysis'
        if row[18].value == "Pending Analysis" or row[18].value == "Pending Ticket Creation" or row[18].value == "Pending Reevaluation":
            for cell in row:
                cell.fill = my_bad
                cell.font = bad_font
        elif row[18].value == "Pending Patch Cycle" or row[18].value == "Pending Remediation" or row[18].value == "On Hold":
            for cell in row:
                cell.fill = my_neutral
                cell.font = neutral_font
        elif re.compile("Remed.*").match(row[18].value) or row[18].value == "Closed":
            for cell in row:
                cell.fill = my_good
                cell.font = good_font
        elif row[18].value == "Risk Ack. Needed" or row[18].value == "False Positive Doc. Needed":
            for cell in row:
                cell.fill = my_check
                cell.font = check_font
                cell.border = gray_border
    return ws

#Generate a fresh workbook for importing vulnerability data
def _Gen_Fresh_Workbook (spreadsheet, sheets):
    statuses_data = {'Status':['Pending Analysis', 'Pending Ticket Creation', # define the data that goes into the default reference sheets
                                'Pending Patch Cycle', 'Pending Remediation', 'Pending Reevaluation',
                                'Risk Ack. Needed', 'False Positive Doc. Needed',
                                'On Hold', 'Closed',
                                'Remediated - Jan', 'Remediated - Feb',
                                'Remediated - Mar', 'Remediated - Apr',
                                'Remediated - May', 'Remediated - Jun',
                                'Remediated - Jul', 'Remediated - Aug',
                                'Remediated - Sep', 'Remediated - Oct',
                                'Remediated - Nov', 'Remediated - Dec'],
                                'Explanation':["The vulnerability has yet to be analyzed by a security analyst",
                                                "The vulnerability has been analyzed by a security analyst, but the assigned analyst has not yet created a ticket for it.",
                                                "The vulnerability is very recent and references one or several critical or important OS patches. This status is used when the analyst determines that the vulnerability does not pose an immediate threat and has no reason to believe the patching procedure will not handle the vulnerability according to its schedule.",
                                                "The vulnerability has been analyzed and a ticket has been created/assigned outlining remediation steps.",
                                                "The vulnerability entry has been modified or marked wrongly in some way, or an oddity in the scan data warrants manual examination. This could be due to a vulnerability's status being marked as \"Remediated\" even though it continues to be detected in scans. Whenever this status is seen, refer to the vulnerability's \"Mitigation Notes\" or \"Robot Notes\" columns for more information.",
                                                "The vulnerability entry has been confirmed to represent a true positive, however certain factors dictate that the vulnerability cannot or will not be remediated and the risk will be accepted instead. This decision usually comes from the Leadership Team or the Head of the Security Department. Every risk acknowledgment needs to be documented according to a template/standard.",
                                                "The vulnerability has been confirmed to represent a false positive. This is usually accompanied by a modified scanner configuration, listing the false positive vulnerability as a lower default risk level to prevent it from appearing as High or Critical. Every false positive needs to be documented according to a template/standard.",
                                                "This status ought not be used frequently. It is a catch-all status for any vulnerability entry that does not meet any of the other status conditions and must be picked back up in the future.",
                                                "This status ought not be used frequently. It is a catch-all status for any vulnerability entry that does not meet any of the other status conditions and does not need to be handled, modified, analyzed, remediated, or otherwise worried about.",
                                                "This status and all other month-based remediation statuses signify the vulnerability has been handled and will not be seen again on the device in question. Having separate remediation statuses for every month allows for the use of macros to generate remediation reports for any given month out of the year.", "\"", "\"", "\"", "\"", "\"", "\"", "\"", "\"", "\"", "\"", "\""]}
    columns_data = {'Column':['Vulnerability Name',
                              'Plugin ID',
                              'Target',
                              'Device Name',
                              'MAC(s)',
                              'OS',
                              'Port',
                              'Service',
                              'Synopsis',
                              'Output',
                              'Last Scanned',
                              'Analysis Date',
                              'Analyst',
                              'Severity',
                              'Risk',
                              'Solution',
                              'Notes',
                              'Ticket #',
                              'Status',
                              'Vulnerability Details',
                              'Scanner Config?',
                              'Robot Note'],
                                'Explanation':["The name of the vulnerability as it is reported by the scan source.",
                                "The scanner plugin that detected the vulnerability.",
                                "The target identifier used by the scan to identify unique targets detected.",
                                "The DNS name or NetBIOS name of the target host - this is not a fixed or globally unique identifier!",
                                "The device's MAC address; contains multiple addresses if the scanner detects multiple network interfaces. If a MAC is not detected, '???' will be listed. This severely limits the Python tool's ability to examine vulnerability entries.",
                                "The detected OS; this is most likely not accurate. Don't rely on it.",
                                "The port that the vulnerability was detected on.",
                                "The service name using the port in question. Determination of the name happens on the scanner's side, not the target's side.",
                                "The basic description of the vulnerability provided by the scanner's plugin description.",
                                "The plugin output that shows you why the plugin picked up the vulnerability on the target host; this is important information for determining the remediation steps required.",
                                "The last time this device was seen in a scan report (does not consider authentication success though)",
                                "The date on which the assigned security analyst performed analysis on the vulnerability",
                                "The security analyst who performed the initial analysis on the vulnerability",
                                "The scanner's reported severity rating for the vulnerability in question; 4 is critical. 3 is High.",
                                "The risk level assigned by the security analyst; this ranking, of critical, high, medium, low, or N/A, signifies the practical threat level - not the abstract default threat level given by Nessus; this ranking informs the remediation prioritization schedule",
                                "The default solution provided by the scanner plugin.",
                                "The security analysts notes on the vulnerability; this could be notes on tracking the vulnerability, more insight into why it was detected, or more details about how to remediate it.",
                                "The ticket number of the ticket created to handle the remediation of the vulnerability",
                                "Describes the stage in which the vulnerability is within the analysis/remediation cycle",
                                "A link to the scanner plugin details; usually a link to an external resource.",
                                "Indicates whether the vulnerability prompted a reconfiguration of the scanner's default plugin threat rankings; provides detail on what the change entails if a change has taken place",
                                "Any notes left by the vuln_analysis.py script for the given vulnerability"]}

    statuses_df = pd.DataFrame(statuses_data) # create the not-yet-formatted dataframes that hold the above data dictionaries
    columns_df = pd.DataFrame(columns_data)

    writer = pd.ExcelWriter(spreadsheet, engine='xlsxwriter') # initialize the new spreadsheet writer object and a workbook object to work with
    workbook = writer.book

    pgraph = workbook.add_format({'text_wrap': True, # define various cell styles and formats compatible with the xlsxwriter writer object
                                    'align': 'left',
                                    'valign': 'top'})
    stat = workbook.add_format({'text_wrap': True,
                                'align': 'left',
                                'valign': 'vcenter'})
    red = workbook.add_format({'bg_color': '#FFC7CE',
                                'font_color': '#9C0006',
                                'align': 'center',
                                'valign': 'vcenter',
                                'border': True,
                                'border_color': 'black'})
    yellow = workbook.add_format({'bg_color': '#FFEB9C',
                                'font_color': '#9C5700',
                                'align': 'left',
                                'valign': 'vcenter',
                                'border': True,
                                'border_color': 'black'})
    gray = workbook.add_format({'bg_color': '#A5A5A5',
                                'font_color': 'white',
                                'bold': True,
                                'align': 'left',
                                'valign': 'vcenter',
                                'border': 6,
                                'border_color': 'black'})
    green = workbook.add_format({'bg_color': '#C6EFCE',
                                'font_color': '#006100',
                                'align': 'left',
                                'valign': 'vcenter',
                                'border': True,
                                'border_color': 'black'})

    for s in sheets: # iterate over the sheetnames you provided previously
        worksheet_df = pd.DataFrame(columns=columns_df['Column']) # create temp dataframe using column names according to those listed in the 'columns' sheet
        worksheet_df.to_excel(writer, s, index=False) # transfer temp dataframe to the spreadsheet object as a new sheet
        ws = writer.sheets[s] # define a new temp sheet object in order to freeze the first row column names and set formats
        ws.freeze_panes(1, 0)

    statuses_df.to_excel(writer, 'statuses', index=False) # transfer default sheet dataframes to the spreadsheet object as new sheets
    columns_df.to_excel(writer, 'columns', index=False)

    statuses_sheet = writer.sheets['statuses'] # initialize worksheet objects out of the previously transferred dataframe sheets in order to apply formatting to them
    columns_sheet = writer.sheets['columns']

    statuses_sheet.set_column(1, 1, 120, pgraph) # format the sheets with appropriate cell widths and styles previously defined
    statuses_sheet.set_column(0, 0, 27, stat)
    columns_sheet.set_column(1, 1, 120, pgraph)
    columns_sheet.set_column(0, 0, 27, stat)
    statuses_sheet.conditional_format('A2:A3', {'type': 'no_errors', 'format': red})
    statuses_sheet.conditional_format('A4:A5', {'type': 'no_errors', 'format': yellow})
    statuses_sheet.conditional_format('A6', {'type': 'no_errors', 'format': red})
    statuses_sheet.conditional_format('A7:A8', {'type': 'no_errors', 'format': gray})
    statuses_sheet.conditional_format('A9', {'type': 'no_errors', 'format': yellow})
    statuses_sheet.conditional_format('A10:A22', {'type': 'no_errors', 'format': green})

    writer.save() # save the spreadsheet object and close it

    wb = load_workbook(spreadsheet, read_only=False) # initialize a new workbook object so we can modify the now-existing xlsx file

    wb.add_named_style(vuln_name_style) # add the previously defined styles to the workbook object so they can be used; this should stick between saves
    wb.add_named_style(the_rest_style)

    for sheet in wb.sheetnames: # iterate over sheetnames as they occur in the spreadsheet previously created and saved
        if sheet in sheets: # check if the current sheet is one of the analysis sheets
            ws1 = wb[sheet] # initialize a worksheet object to apply styles and widths
            ws1.add_data_validation(data_val) # applies data validation to the statuses column so that the program's modification logic doesn't hit any snags
            data_val.add("S2:S1048576")
            ws1 = _Set_Col_Styles(ws1) # iterate over cells in specified columns and apply styles
            ws1 = _Set_Col_Widths(ws1) # set custom column widths
    # finally save and close the fresh worksheet, ready to be fed into the program
    wb.save(spreadsheet)

# Modifies existing entries in the target sheet only based on vulnerabilities found (or not found) in the new report
def _Mod_Analysis_Spreadsheet (vuln_analysis_df, report_df, report_dict):
    common_name_indices = []
    common_indices = []
    report_indices = []
    diff_indices = []
    # check each existing vulnerability in the analysis spreadsheet for matches in both the vuln name AND mac address columns in the new report dataframe; then modify the existing vulnerability row in-place based on various checks
    for old_vuln in vuln_analysis_df['Vulnerability Name']:
        lst = report_df[report_df['Vulnerability Name'] == old_vuln].index.tolist()
        common_name_indices.append(lst) # list of lists of indices specifying where the spreadsheet contains vuln names in common with the report vuln names
    for lst_no in range(len(common_name_indices)):
        for index in common_name_indices[lst_no]:
            if vuln_analysis_df.iloc[lst_no]['MAC(s)'] == report_df.iloc[index]['MAC(s)']:
                if report_df.iloc[index]['MAC(s)'] == '???':
                    vuln_analysis_df.loc[(lst_no, 'Status')] = 'Pending Reevaluation'
                    vuln_analysis_df.loc[(lst_no, 'Robot Note')] = 'A MAC address could not be detected for this device, but it was in a recent scan - please manually determine the status of this vulnerability (delete this note)'
                else:
                    common_indices.append(lst_no) # pick out list of list indices where BOTH the vuln names and mac address cells represent matches between both dataframes
                    report_indices.append(index)

    for row in common_indices: # iterate over rows (that represent vulnerabilities in common with the new report) that need to be checked pending modification
        if report_dict[vuln_analysis_df.iloc[row]['Target']]['auth'] == 's':
            if datetime.datetime.strptime(vuln_analysis_df.iloc[row]['Last Scanned'], '%a %b %d %H:%M:%S %Y') <= datetime.datetime.strptime(report_dict[vuln_analysis_df.iloc[row]['Target']]['HOST_START'], '%a %b %d %H:%M:%S %Y'):
                vuln_analysis_df.loc[(row, 'Last Scanned')] = report_dict[vuln_analysis_df.iloc[row]['Target']]['HOST_START'] # always change the Last Scanned cell to the report's scan date (as long as the report is, in fact, newer)
                if vuln_analysis_df.iloc[row]['Status'] == 'Pending Patch Cycle': # change the status of rows needing a reevaluation based on patch cycle
                    vuln_analysis_df.loc[(row, 'Status')] = 'Pending Reevaluation'
                    vuln_analysis_df.loc[(row, 'Robot Note')] = 'was pending patch cycle - re-examine vulnerability.'
                elif vuln_analysis_df.iloc[row]['Status'] == 'Pending Ticket Creation': # change the status of rows needing reevaluation based on whether risk was low and remediation was delayed
                    if vuln_analysis_df.iloc[row]['Risk'] == 'Med' or vuln_analysis_df.iloc[row]['Risk'] == 'Low':
                        vuln_analysis_df.loc[(row, 'Status')] = 'Pending Reevaluation'
                        vuln_analysis_df.loc[(row, 'Robot Note')] = 'was pending ticket creation and med/low risk - time to process it now?'
                    if vuln_analysis_df.iloc[row]['Risk'] == 'High' or vuln_analysis_df.iloc[row]['Risk'] == 'Crit':
                        vuln_analysis_df.loc[(row, 'Status')] = 'Pending Reevaluation'
                        vuln_analysis_df.loc[(row, 'Robot Note')] = 'was pending ticket creation and crit/high risk - HANDLE IT THIS CYCLE.'
                elif re.compile("Remed.*").match(vuln_analysis_df.iloc[row]['Status']): # change the status of rows that were marked remediated in error
                    vuln_analysis_df.loc[(row, 'Status')] = 'Pending Reevaluation'
                    vuln_analysis_df.loc[(row, 'Robot Note')] = 'marked remediated but was picked up in last scan - re-examine host.'
        if vuln_analysis_df.iloc[row]['Status'] == None: # catch any rows with empty Status cells (there should never be rows with empty Status cells)
            vuln_analysis_df.loc[(row, 'Status')] = 'Pending Analysis'

    for i in range(len(vuln_analysis_df['Vulnerability Name'])): # build list of individual spreadsheet indices that represent vulnerabilities that do not appear in the provided scan report
        if i not in common_indices:
            diff_indices.append(i)

    for row in diff_indices: # iterate over rows (that represent vulnerabilities that do not reappear in the provided scan report) that need to be checked pending modification
        if report_dict[vuln_analysis_df.iloc[row]['Target']]['auth'] == 's':
            if datetime.datetime.strptime(vuln_analysis_df.iloc[row]['Last Scanned'], '%a %b %d %H:%M:%S %Y') <= datetime.datetime.strptime(report_dict[vuln_analysis_df.iloc[row]['Target']]['HOST_START'], '%a %b %d %H:%M:%S %Y'):
                vuln_analysis_df.loc[(row, 'Last Scanned')] = report_dict[vuln_analysis_df.iloc[row]['Target']]['HOST_START'] # always change the Last Scanned cell to the report's scan date (as long as the report is, in fact, newer)
                if (vuln_analysis_df.iloc[row]['Status'] == 'Pending Remediation' or vuln_analysis_df.iloc[row]['Status'] == 'Pending Ticket Creation' or vuln_analysis_df.iloc[row]['Status'] == 'Pending Analysis' or vuln_analysis_df.iloc[row]['Status'] == 'Pending Patch Cycle'):
                    vuln_analysis_df.loc[(row, 'Status')] = 'Remediated'+' - '+(DATE.strftime('%b'))
                    vuln_analysis_df.loc[(row, 'Robot Note')] = 'was pending, and was not found in the last credentialed check of the host - marked remediated.'
        if vuln_analysis_df.iloc[row]['Status'] == None: # catch any rows with empty Status cells (there should never be rows with empty Status cells)
            vuln_analysis_df.loc[(row, 'Status')] = 'Pending Analysis'

# Identifies never-before-seen vulnerabilities in the report dataframe and appends them to the spreadsheet dataframe
def _Add_New_Vulns (vuln_analysis_df, report_df):
    merged_df = report_df.merge(vuln_analysis_df, how='inner', on = ['Vulnerability Name', 'Device Name', 'MAC(s)'], suffixes=('','_y')) # generate a dataframe with rows that match between the sheet df and the report df
    merged_df.drop(list(merged_df.filter(regex='_y$')), axis=1, inplace=True) # strip away unwanted columns created by the merge
    diff_df = pd.concat([report_df, merged_df], sort=False) # concatenate the report df and the df containg similarities between the sheet df and the report df
    diff_df2 = diff_df.drop_duplicates(subset=['Vulnerability Name', 'Target', 'MAC(s)'],keep=False) # drop all except unique entries, leaving us only with report df vulnerability/host combos that are totally unique to the report and never appear in the sheet df
    return vuln_analysis_df.append(diff_df2, ignore_index=True, sort=False) # return the generated final df onto the working sheet df

# Performs all modification of the analysis spreadsheet after analyzing the scan reports
def _Finagle_WB (existing_spreadsheet, wb, vuln_analysis_df, target_sheet):
    writer = pd.ExcelWriter(existing_spreadsheet, engine='openpyxl') # declare engine to write dataframes to the spreadsheet
    writer.book = wb # define the workbook that the writer writes to
    print("Making changes in "+existing_spreadsheet.split('\\')[-1]+"...")
    wb.remove(wb[target_sheet]) # remove the unedited sheet in prep for adding modified ones

    vuln_analysis_df.to_excel(writer, sheet_name=target_sheet, index=False, engine='openpyxl') # delicately place new dataframes into the excel spreadsheet and define a new worksheet object to add in-place formatting to
    ws1 = wb[target_sheet]

    ws1 = _Set_Col_Styles(ws1) # apply baseline alignment and border formats to appropriate columns in both the working sheet and targets sheet

    ws1 = _Set_Row_Format(ws1) # fine-tune formatting (color, border, font, etc.) based on vulnerability status

    sheetnames = [s for s in wb.sheetnames if s != 'statuses' or s != 'columns'] # get list of sheets to iterate through so we can apply data validation to Statuses columns

    for s in sheetnames: # apply dropdown menu data validation to the Statuses column in every sheet
        for sheet in wb.sheetnames:
            if s == sheet:
                ws2 = wb[s]
                ws2.add_data_validation(data_val)
                data_val.add("S2:S1048576") # specifies the column/rows to apply to; the second value means ALL rows under column O

    ws1 = _Set_Col_Widths(ws1) # set adequate column widths for all columns in the working sheet as well as the targets sheet

    ws1.freeze_panes = "A2" # freeze top row column names

    print("Saving and closing "+existing_spreadsheet.split('\\')[-1]+".")
    # save and close objects, finalizing spreadsheet changes
    wb.save(existing_spreadsheet)

# Function to run if user chose '1'
def _1_Create_Fresh_Spreadsheet (spreadsheet, sheets):
    if spreadsheet == '':
        spreadsheet = _Check_Path(input("Enter full path and name for your new spreadsheet: "), 'v')
    else:
        spreadsheet = _Check_Path(spreadsheet, 'v')
    if sheets == '':
        sheetz = input("Enter a comma-separated list (no spaces) of sheet names to create: ").replace(" ", "").split(',')
    else:
        sheetz = sheets.strip(" ").split(',')
    _Gen_Fresh_Workbook(spreadsheet, sheetz)

# Function to run if user chose '2'
def _2_Feed_New_Reports (nessusfile, spreadsheet, sheet):
    if nessusfile == '':
        nessusfile = _Check_Path(input("Enter a filepath to your .nessus file: "), 'n') # Provide path to .nessus report file for importing
    else:
        nessusfile = _Check_Path(nessusfile, 'n')

    report_dict, client = _Parse_Nessus(nessusfile) # parse .nessus file for report dict and sheet (client) name

    if spreadsheet == '':
        spreadsheet = _Check_Path(input("Enter a path to your existing analysis spreadsheet"), 'x')
    else:
        spreadsheet = _Check_Path(spreadsheet, 'x')

    _Backup(spreadsheet)
    wb = load_workbook(spreadsheet, read_only=False)

    try:
        ws = wb[client]
        print("Target sheet name automatically gathered according to Scan name.")
        sheet = client
    except:
        if sheet == '':
            sheet = _Check_Sheet(input("Enter the name of the worksheet to load the results into: "), wb)
        else:
            sheet = _Check_Sheet(sheet, wb)

        ws = wb[sheet]
    data = ws.values # for defining dataframe contents
    columns = next(data)[0:] # for defining dataframe column names

    print("Initializing and preparing vulnerability dataframes...\n")
    vuln_analysis_df = pd.DataFrame(data, columns=columns)
    vuln_analysis_df.dropna(axis=0, how='all', inplace=True) # drop null rows and clear the index away totally
    report_df = pd.DataFrame().reindex_like(vuln_analysis_df) # create an empty duplicate of spreadsheet df for working with the Nessus reports

    print("Building report dataframe...")
    row = 0
    for target in report_dict:
        for v in report_dict[target]['vulns']:
            if report_dict[target]['vulns'][v]['severity'] == '3' or report_dict[target]['vulns'][v]['severity'] == '4':
                report_df.loc[row, 'Vulnerability Name'] = report_dict[target]['vulns'][v]['pluginName']
                report_df.loc[row, 'Plugin ID'] = report_dict[target]['vulns'][v]['pluginID']
                report_df.loc[row, 'Target'] = target
                # import Device Name
                if 'host-rdns' in report_dict[target].keys():
                    report_df.loc[row, 'Device Name'] = report_dict[target]['host-rdns']
                elif 'netbios-name' in report_dict[target].keys():
                    report_df.loc[row, 'Device Name'] = report_dict[target]['netbios-name']
                else:
                    report_df.loc[row, 'Device Name'] = report_dict[target]['host-ip']
                # import MAC(s)
                if 'mac-address' in report_dict[target].keys():
                    report_df.loc[row, 'MAC(s)'] = report_dict[target]['mac-address']
                else:
                    report_df.loc[row, 'MAC(s)'] = '???'
                report_df.loc[row, 'OS'] = report_dict[target]['operating-system']
                report_df.loc[row, 'Port'] = report_dict[target]['vulns'][v]['port']
                report_df.loc[row, 'Service'] = report_dict[target]['vulns'][v]['svc_name']
                report_df.loc[row, 'Synopsis'] = report_dict[target]['vulns'][v]['synopsis']
                if 'plugin_output' in report_dict[target]['vulns'][v].keys():
                    report_df.loc[row, 'Output'] = report_dict[target]['vulns'][v]['plugin_output']
                else:
                    report_df.loc[row, 'Output'] = 'N/A'
                report_df.loc[row, 'Last Scanned'] = report_dict[target]['HOST_START'] # EX: datetime.datetime.strptime('Tue Jan 26 08:56:53 2021', '%a %b %d %H:%M:%S %Y')
                report_df.loc[row, 'Severity'] = report_dict[target]['vulns'][v]['severity']
                report_df.loc[row, 'Solution'] = report_dict[target]['vulns'][v]['solution']
                report_df.loc[row, 'Vulnerability Details'] = 'https://www.tenable.com/plugins/nessus/' + report_dict[target]['vulns'][v]['pluginID']
                row+=1

    print("Modifying target analysis sheet with new scan data...")
    _Mod_Analysis_Spreadsheet(vuln_analysis_df, report_df, report_dict) # change the existing spreadsheet's dataframe to reflect new report data
    vuln_analysis_df = _Add_New_Vulns(vuln_analysis_df, report_df) # add new vulnerability/target combos to the analysis dataframe
    _Finagle_WB(spreadsheet, wb, vuln_analysis_df, sheet)

# Function to run if user chose '3'
def _3_Add_New_Sheet (spreadsheet, new_sheet):
    if spreadsheet == '':
        spreadsheet = _Check_Path(input("Enter a path to your existing analysis spreadsheet"), 'x')
    else:
        spreadsheet = _Check_Path(spreadsheet, 'x')

    _Backup(spreadsheet)
    wb = load_workbook(spreadsheet, read_only=False)
    sheets = [s for s in wb.sheetnames if s != 'columns' and s != 'statuses']
    ws = wb[sheets[0]] # get a sheet to duplicate
    data = ws.values # for defining dataframe contents
    columns = next(data)[0:] # for defining dataframe column names
    df = pd.DataFrame(data, columns=columns)
    df.dropna(axis=0, how='all', inplace=True)
    new_df = pd.DataFrame().reindex_like(df)

    if new_sheet == '':
        new_sheet = input("Enter the name of the new sheet (only one will be added): ")

    writer = pd.ExcelWriter(spreadsheet, engine='openpyxl')
    writer.book = wb
    new_df.to_excel(writer, sheet_name=new_sheet, index=False, engine='openpyxl')
    ws1 = wb[new_sheet]

    ws1.add_data_validation(data_val) # applies data validation to the statuses column so that the program's modify logic doesn't hit any snags
    data_val.add('S2:S1048576')
    ws1 = _Set_Col_Styles(ws1) # iterate over cells in specified columns and apply styles
    ws1 = _Set_Col_Widths(ws1) # set column widths
    ws1.freeze_panes = "A2" # freeze top row column names

    wb.save(spreadsheet) # finally save and close the workbook

# Function to run if user chose '4'
def _4_Generate_Remed_Report (spreadsheet, sheet, month):
    if spreadsheet == '':
        spreadsheet = _Check_Path(input("Enter a path to your existing analysis spreadsheet"), 'x')
    else:
        spreadsheet = _Check_Path(spreadsheet, 'x')

    wb = load_workbook(spreadsheet, read_only=False) # create a workbook object for writing

    if sheet == '':
        sheet = _Check_Sheet(input("Enter the name of the worksheet to load remediations from: "), wb)
    else:
        sheet = _Check_Sheet(new_sheet, wb)

    while True:
        if month == '':
            print("----------Choose a month to generate report for----------")
            print("--------- 1  - January")
            print("--------- 2  - February")
            print("--------- 3  - March")
            print("--------- 4  - April")
            print("--------- 5  - May")
            print("--------- 6  - June")
            print("--------- 7  - July")
            print("--------- 8  - August")
            print("--------- 9  - September")
            print("--------- 10 - October")
            print("--------- 11 - November")
            print("--------- 12 - December")
            month = input("Enter Number 0-12: ")

        month = int(month)
        if 0 < month <= 12:
            if month == 1:
                mo = 'Remediated - Jan'
            elif month == 2:
                mo = 'Remediated - Feb'
            elif month == 3:
                mo = 'Remediated - Mar'
            elif month == 4:
                mo = 'Remediated - Apr'
            elif month == 5:
                mo = 'Remediated - May'
            elif month == 6:
                mo = 'Remediated - Jun'
            elif month == 7:
                mo = 'Remediated - Jul'
            elif month == 8:
                mo = 'Remediated - Aug'
            elif month == 9:
                mo = 'Remediated - Sep'
            elif month == 10:
                mo = 'Remediated - Oct'
            elif month == 11:
                mo = 'Remediated - Nov'
            elif month == 12:
                mo = 'Remediated - Dec'
            break
        else:
            print("Invalid Number")
            month = ''
            continue

    ws = wb[sheet]
    report = Workbook() # create a new excel file to be the remediation report
    fn = '\\'.join(spreadsheet.split('\\')[:-1])+'\\'+sheet+' Remediation Report_'+mo.split()[-1]+'.xlsx' # the name of the report file
    rws = report.active
    rws.title = sheet+' Remed. Report'
    rws.merge_cells('A1:V1')
    rws['A1'] = sheet+' Remediation Report - '+mo.split()[-1]
    report.save(filename = fn)
    report.close()

    data = ws.values
    columns = next(data)[0:]

    vuln_analysis_df = pd.DataFrame(data, columns=columns) # define and manipulate dataframes in preparation for working with them
    vuln_analysis_df.dropna(axis=0, how='all', inplace=True) # must do this to drop null rows and clear the index away totally
    remed_df = pd.DataFrame().reindex_like(vuln_analysis_df) # create an empty duplicate of spreadsheet df for collecting remediated rows
    remed_df.dropna(axis=0, how='all', inplace=True)

    indices = []
    for i in range(len(vuln_analysis_df['Status'])): # iterate over statuses in each row and determine if each one corresponds to the requested month
        if vuln_analysis_df.loc[(i, 'Status')] == mo:
            indices.append(i) # save the index of the match to the list

    for index in indices: # add the correct rows to the report excel file
        remed_df = remed_df.append(vuln_analysis_df.loc[index], ignore_index=True, sort=False)

    report_wb = load_workbook(fn) # the rest is all formatting the report excel file
    wsr = report_wb.active
    for r in dataframe_to_rows(remed_df, index=False, header=True):
        wsr.append(r)

    report_wb.add_named_style(vuln_name_style)
    report_wb.add_named_style(the_rest_style)

    wsr = _Set_Col_Styles(wsr)

    wsr = _Set_Col_Widths(wsr) # set column widths

    wsr['A1'].alignment = Alignment(horizontal='center', vertical='center')
    wsr['A1'].font = Font(color='FFFFFF', bold=True, size='18')
    wsr['A1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    wsr.freeze_panes = "A3"

    for row in wsr.iter_rows():
        if row[18].value != None:
            if re.compile("Remed.*").match(row[18].value):
                for cell in row:
                    cell.fill = my_good
                    cell.font = good_font

    report_wb.save(fn)
    print("\nYour report has been saved in the same directory as the spreadsheet.\n")

# Function to run if user chose '5'
def _5_Migrate_Spreadsheet (spreadsheet):
    if spreadsheet == '':
        spreadsheet = _Check_Path(input("Enter a path to your existing analysis spreadsheet"), 'x')
    else:
        spreadsheet = _Check_Path(spreadsheet, 'x')

    new_spreadsheet = _Check_Path(input("Enter full path and name for your new spreadsheet: "), 'v')
    wb = load_workbook(spreadsheet) # load the spreadhseet to translate to a new version
    sheets = [s for s in wb.sheetnames if s != 'columns' and s != 'statuses']
    _Gen_Fresh_Workbook(new_spreadsheet, sheets)
    wb2 = load_workbook(new_spreadsheet)
    writer = pd.ExcelWriter(new_spreadsheet, engine='openpyxl')
    writer.book = wb2

    for s in sheets: # determine vulns that are still active and in question - exclude vulns that have been remediated or closed
        ws = wb[s]
        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
        df2 = df.loc[~((df.Status.str.match('Remed.*')) | (df['Status']=='Closed'))]
        df2.dropna(axis=0, how='all', inplace=True)
        wb2.remove(wb2[s])
        df2.to_excel(writer, sheet_name=s, index=False, engine='openpyxl')
        ws2 = wb2[s]

        ws2 = _Set_Col_Styles(ws2)

        ws2.add_data_validation(data_val)
        data_val.add('S2:S1048576')

        ws2 = _Set_Row_Format(ws2) # fine-tune formatting (color, border, font, etc.) based on vulnerability status

        ws2 = _Set_Col_Widths(ws2)

        ws2.freeze_panes = "A2" # freeze top row column names

        wb2.save(new_spreadsheet)

def _Cycle_Opts (opts):
    nessusfile = ''
    spreadsheet = ''
    sheets = ''
    month = ''
    for opt, arg in opts:
        if opt == "-n":
            nessusfile = arg
        elif opt == "-s":
            spreadsheet = arg
        elif opt == "-t":
            sheets = arg
        elif opt == "-m":
            month = arg
    return nessusfile, spreadsheet, sheets, month

def main (argv):
    try:
        opts, args = getopt.getopt(argv,"hi12345n:s:t:m:",["nessusfile=","spreadsheet=","sheets=","month="])
    except getopt.GetoptError:
        _Opt_Help()
        exit(2)

    selection = 0
    while selection == 0:
        for opt, arg in opts:
            if opt == '-h':
                _Opt_Help()
                exit()
            elif opt == "-i":
                break
            elif opt == "-1":
                selection = 1
            elif opt == "-2":
                selection = 2
            elif opt == "-3":
                selection = 3
            elif opt == "-4":
                selection = 4
            elif opt == "-5":
                selection = 5

        if selection == 0:
            print("----------MENU----------")
            print("1. Create fresh spreadsheet")
            print("2. Feed new reports")
            print("3. Add a new sheet to existing spreadsheet")
            print("4. Generate a Remediation Report")
            print("5. Transition to new workbook")
            print("6. Exit")
            selection = int(input("Enter a number option: "))

    nessusfile, spreadsheet, sheets, month = _Cycle_Opts(opts)
    if selection == 1:
        _1_Create_Fresh_Spreadsheet(spreadsheet, sheets)
        exit()
    if selection == 2:
        _2_Feed_New_Reports(nessusfile, spreadsheet, sheets)
        exit()
    if selection == 3:
        _3_Add_New_Sheet(spreadsheet, sheets)
        exit()
    if selection == 4:
        _4_Generate_Remed_Report(spreadsheet, sheets, month)
        exit()
    if selection == 5:
        _5_Migrate_Spreadsheet(spreadsheet)
        exit()
    if selection == 6:
        exit()

main(argv[1:])
